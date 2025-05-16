# data_processing/amdec_generator.py
import pandas as pd
import numpy as np
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import random

class AMDECGenerator:
    """
    Classe pour générer une analyse AMDEC à partir des données d'historique
    """
    
    def __init__(self, df):
        """
        Initialisation avec les données d'historique
        
        Args:
            df (pandas.DataFrame): DataFrame contenant les données d'historique
        """
        self.df = df
        self.amdec_data = []
        self.amdec_df = None
        self.output_path = None
        
        # Vérifier si les colonnes requises sont présentes
        required_columns = ['composant', 'sous_composant', 'cause', 'duree']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise ValueError(f"Colonnes manquantes dans les données: {', '.join(missing_columns)}")
    
    def generate(self):
        """
        Génère l'analyse AMDEC à partir des données d'historique
        """
        # Grouper les données par composant et sous-composant
        grouped = self.df.groupby(['composant', 'sous_composant'])
        
        for (component, subcomponent), group in grouped:
            # Ignorer les composants/sous-composants inconnus
            if component == 'Inconnu' or subcomponent == 'Inconnu':
                continue
            
            # Pour chaque cause unique, calculer la fréquence, gravité et détection
            causes = group['cause'].value_counts().reset_index()
            causes.columns = ['cause', 'frequency']
            
            for _, row in causes.iterrows():
                cause = row['cause']
                
                # Récupérer les arrêts liés à cette cause spécifique
                related_stops = group[group['cause'] == cause]
                
                # Calculer la fréquence (F) - basée sur le nombre d'occurrences
                frequency_count = len(related_stops)
                frequency_value = self._calculate_frequency(frequency_count)
                
                # Calculer la gravité (G) - basée sur la durée moyenne des arrêts
                avg_duration = related_stops['duree'].mean()
                gravity_value = self._calculate_gravity(avg_duration)
                
                # Calculer la détection (D) - basée sur une heuristique simple
                detection_value = self._calculate_detection(cause)
                
                # Calculer la criticité (C)
                criticality = frequency_value * gravity_value * detection_value
                
                # Déterminer le mode de défaillance
                failure_mode = self._determine_failure_mode(cause, subcomponent)
                
                # Déterminer l'effet
                effect = self._determine_effect(cause, failure_mode)
                
                # Générer des actions correctives
                corrective_actions = self._generate_corrective_actions(component, subcomponent, failure_mode, cause, criticality)
                
                # Ajouter à l'AMDEC
                self.amdec_data.append({
                    'Composant': component.title(),
                    'Sous-composant': subcomponent.title(),
                    'Fonction': self._determine_function(component, subcomponent),
                    'Mode de Défaillance': failure_mode,
                    'Cause': cause.title(),
                    'Effet': effect,
                    'F': frequency_value,
                    'G': gravity_value,
                    'D': detection_value,
                    'C': criticality,
                    'Actions Correctives': corrective_actions
                })
        
        # Convertir en DataFrame
        if self.amdec_data:
            self.amdec_df = pd.DataFrame(self.amdec_data)
            
            # Trier par composant, puis par criticité (descendant)
            self.amdec_df = self.amdec_df.sort_values(by=['Composant', 'C'], ascending=[True, False])
        else:
            # Créer un DataFrame vide avec les bonnes colonnes
            self.amdec_df = pd.DataFrame(columns=[
                'Composant', 'Sous-composant', 'Fonction', 'Mode de Défaillance', 
                'Cause', 'Effet', 'F', 'G', 'D', 'C', 'Actions Correctives'
            ])
        
        return self.amdec_df
    
    def _calculate_frequency(self, count):
        """
        Calcule la valeur de fréquence (F) pour l'AMDEC
        
        Args:
            count (int): Nombre d'occurrences
            
        Returns:
            int: Valeur de fréquence entre 1 et 4
        """
        if count <= 1:
            return 1  # Défaillance rare
        elif count <= 3:
            return 2  # Défaillance possible
        elif count <= 6:
            return 3  # Défaillance fréquente
        else:
            return 4  # Défaillance très fréquente
    
    def _calculate_gravity(self, duration):
        """
        Calcule la valeur de gravité (G) pour l'AMDEC basée sur la durée
        
        Args:
            duration (float): Durée moyenne des arrêts (en heures)
            
        Returns:
            int: Valeur de gravité entre 1 et 5
        """
        if duration <= 0.5:  # 30 minutes
            return 1  # Mineure
        elif duration <= 1:  # 1 heure
            return 2  # Significative
        elif duration <= 5:  # 5 heures
            return 3  # Moyenne
        elif duration <= 12:  # 12 heures
            return 4  # Majeure
        else:
            return 5  # Catastrophique
    
    def _calculate_detection(self, cause):
        """
        Calcule la valeur de détection (D) pour l'AMDEC
        
        Args:
            cause (str): Cause de la défaillance
            
        Returns:
            int: Valeur de détection entre 1 et 4
        """
        # Causes facilement détectables
        easily_detectable = ['fuite', 'percement', 'surchauffe', 'vibration']
        
        # Causes moyennement détectables
        medium_detectable = ['corrosion', 'erosion', 'encrassement', 'mauvais montage']
        
        # Causes difficilement détectables
        hard_detectable = ['fissure', 'fatigue']
        
        cause = cause.lower()
        
        if any(ed in cause for ed in easily_detectable):
            return 1  # Détection évidente
        elif any(md in cause for md in medium_detectable):
            return 2  # Détection possible
        elif any(hd in cause for hd in hard_detectable):
            return 3  # Détection improbable
        else:
            return 4  # Détection impossible
    
    def _determine_failure_mode(self, cause, subcomponent):
        """
        Détermine le mode de défaillance en fonction de la cause et du sous-composant
        
        Args:
            cause (str): Cause de la défaillance
            subcomponent (str): Sous-composant concerné
            
        Returns:
            str: Mode de défaillance
        """
        cause = cause.lower()
        
        # Mappings des causes aux modes de défaillance
        failure_mode_mappings = {
            'corrosion': {
                'epingle': 'Corrosion externe',
                'collecteur': 'Corrosion interne',
                'tube': 'Corrosion côté feu',
                'branches': 'Corrosion sous contrainte'
            },
            'fissure': {
                'epingle': 'Fissuration thermique',
                'collecteur': 'Fissures intergranulaires',
                'tube': 'Fatigue thermique',
                'branches': 'Fissuration'
            },
            'erosion': {
                'epingle': 'Erosion par cendres',
                'collecteur': 'Erosion',
                'tube': 'Erosion par particules',
                'branches': 'Erosion'
            },
            'fatigue': {
                'epingle': 'Fatigue thermique',
                'collecteur': 'Fatigue mécanique',
                'tube': 'Fatigue',
                'branches': 'Fatigue cyclique'
            },
            'percement': {
                'epingle': 'Percement par érosion',
                'collecteur': 'Fuites locales',
                'tube': 'Percement',
                'branches': 'Rupture par fluage'
            },
            'surchauffe': {
                'epingle': 'Short-term overheat',
                'collecteur': 'Long-term overheat',
                'tube': 'Surchauffe locale',
                'branches': 'Surchauffe'
            },
            'encrassement': {
                'epingle': 'Encrassement interne',
                'collecteur': 'Dépôts',
                'tube': 'Bouchage',
                'branches': 'Encrassement'
            },
            'vibration': {
                'epingle': 'Vibration excessive',
                'collecteur': 'Vibration',
                'tube': 'Vibrations du flux',
                'branches': 'Vibration des branches'
            },
            'fuite': {
                'epingle': 'Fuite',
                'collecteur': 'Fuite aux joints',
                'tube': 'Fuite',
                'branches': 'Fuite aux raccords'
            }
        }
        
        # Recherche de correspondance dans les clés de causes
        for cause_key, subcmp_dict in failure_mode_mappings.items():
            if cause_key in cause:
                # Recherche de correspondance dans les sous-composants
                for subcmp_key, failure_mode in subcmp_dict.items():
                    if subcmp_key in subcomponent.lower():
                        return failure_mode
        
        # Si aucune correspondance n'est trouvée, retourner une valeur par défaut
        return "Défaillance"
    
    def _determine_effect(self, cause, failure_mode):
        """
        Détermine l'effet de la défaillance
        
        Args:
            cause (str): Cause de la défaillance
            failure_mode (str): Mode de défaillance
            
        Returns:
            str: Effet de la défaillance
        """
        # Mappings des modes de défaillance aux effets
        effects = {
            'Corrosion externe': 'Amincissement parois',
            'Corrosion interne': 'Perte matière interne',
            'Corrosion côté feu': 'Perte métal externe',
            'Corrosion sous contrainte': 'Fissures intergranulaires',
            'Fissuration thermique': 'Fissures externes',
            'Fissures intergranulaires': 'Rupture soudure',
            'Fatigue thermique': 'Fissures',
            'Fissuration': 'Fuites locales',
            'Erosion par cendres': 'Amincissement accéléré',
            'Erosion': 'Perte matière',
            'Erosion par particules': 'Surface "fromage suisse"',
            'Fatigue mécanique': 'Fissures externes',
            'Fatigue': 'Rupture',
            'Fatigue cyclique': 'Microfissures',
            'Percement par érosion': 'Perte fluide',
            'Fuites locales': 'Perte rendement',
            'Percement': 'Arrêt d\'urgence',
            'Rupture par fluage': 'Défaillance catastrophique',
            'Short-term overheat': 'Rupture ductile',
            'Long-term overheat': 'Rupture fluage',
            'Surchauffe locale': 'Déformation permanente',
            'Surchauffe': 'Fragilisation',
            'Encrassement interne': 'Réduction débit',
            'Dépôts': 'Blocage flux vapeur',
            'Bouchage': 'Réduction efficacité',
            'Encrassement': 'Surchauffe locale',
            'Vibration excessive': 'Fatigue accélérée',
            'Vibration': 'Desserrage fixations',
            'Vibrations du flux': 'Usure prématurée',
            'Vibration des branches': 'Rupture aux raccords',
            'Fuite': 'Perte rendement',
            'Fuite aux joints': 'Arrêt pour maintenance',
            'Fuite aux raccords': 'Contamination environnement'
        }
        
        # Retourner l'effet correspondant au mode de défaillance
        return effects.get(failure_mode, 'Impact sur performance')
    
    def _determine_function(self, component, subcomponent):
        """
        Détermine la fonction du sous-composant
        
        Args:
            component (str): Composant principal
            subcomponent (str): Sous-composant
            
        Returns:
            str: Fonction du sous-composant
        """
        # Mappings des sous-composants aux fonctions
        functions = {
            'economiseur bt': {
                'epingle': 'Transfert thermique',
                'collecteur entree': 'Alimentation eau',
                'collecteur sortie': 'Collecte eau chauffée',
                'branches entree': 'Distribution eau',
                'branches sortie': 'Évacuation eau',
                'tube porteur': 'Support structurel',
                'tubes suspension': 'Support mécanique'
            },
            'economiseur ht': {
                'epingle': 'Transfert thermique',
                'collecteur entree': 'Distribution vapeur',
                'collecteur sortie': 'Collecte vapeur',
                'branches entree': 'Amenée fluide',
                'branches sortie': 'Évacuation vapeur',
                'tube porteur': 'Support structurel',
                'tubes suspension': 'Support mécanique'
            },
            'surchauffeur bt': {
                'epingle': 'Transfert thermique',
                'collecteur entree': 'Distribution vapeur',
                'collecteur sortie': 'Collecte vapeur surchauffée',
                'tube porteur': 'Résistance pression',
                'branches entree': 'Alimentation vapeur',
                'branches sortie': 'Évacuation vapeur'
            },
            'surchauffeur ht': {
                'epingle': 'Transfert thermique',
                'collecteur entree': 'Distribution vapeur',
                'collecteur sortie': 'Évacuation vapeur',
                'tube porteur': 'Stabilité mécanique',
                'branches entree': 'Distribution vapeur',
                'branches sortie': 'Collecte vapeur'
            },
            'rechauffeur bt': {
                'epingle': 'Échange thermique',
                'collecteur entree': 'Distribution vapeur',
                'collecteur sortie': 'Collecte vapeur',
                'tube porteur': 'Support mécanique',
                'tubes suspension': 'Support structurel',
                'branches entree': 'Alimentation vapeur',
                'branches sortie': 'Évacuation vapeur'
            },
            'rechauffeur ht': {
                'epingle': 'Échange thermique',
                'collecteur entree': 'Distribution vapeur',
                'collecteur sortie': 'Évacuation vapeur',
                'tube porteur': 'Résistance pression',
                'branches entree': 'Amenée vapeur',
                'branches sortie': 'Évacuation vapeur'
            }
        }
        
        # Normaliser les noms pour la recherche
        component = component.lower()
        subcomponent = subcomponent.lower()
        
        # Retourner la fonction correspondante
        if component in functions and subcomponent in functions[component]:
            return functions[component][subcomponent]
        
        # Valeur par défaut si aucune correspondance n'est trouvée
        return "Support et échange thermique"
    
    def _generate_corrective_actions(self, component, subcomponent, failure_mode, cause, criticality):
        """
        Génère des actions correctives en fonction des paramètres
        
        Args:
            component (str): Composant principal
            subcomponent (str): Sous-composant
            failure_mode (str): Mode de défaillance
            cause (str): Cause de la défaillance
            criticality (int): Criticité calculée
            
        Returns:
            str: Actions correctives recommandées
        """
        # Mappings de base des actions correctives par cause
        base_actions = {
            'corrosion': [
                'Revêtement céramique',
                'Contrôle humidité hebdo',
                'Application inhibiteur de corrosion',
                'Analyse chimique eau',
                'Passivation annuelle',
                'Injection additifs anti-corrosion'
            ],
            'fissure': [
                'Inspection PAUT annuelle',
                'Soudure inox',
                'Contrôle ultrasons',
                'Surveillance des contraintes',
                'Test ressuage annuel',
                'Inspection thermique'
            ],
            'erosion': [
                'Revêtement dur',
                'Contrôle particules',
                'Nettoyage pneumatique mensuel',
                'Installation de déflecteurs',
                'Inspection épaisseur par ultrasons',
                'Analyse des cendres'
            ],
            'fatigue': [
                'Renforts métalliques',
                'Surveillance vibratoire',
                'Analyse de contraintes',
                'Modification des supports',
                'Optimisation de la distribution',
                'Inspection par courants de Foucault'
            ],
            'percement': [
                'Remplacement préventif',
                'Contrôle épaisseur mensuel',
                'Revêtement protecteur',
                'Analyse des causes',
                'Optimisation des paramètres',
                'Adaptation des soudures'
            ],
            'surchauffe': [
                'Alarmes température',
                'Purge démarrage',
                'Optimisation combustion',
                'Capteurs température',
                'Contrôle débit',
                'Équilibrage thermique'
            ],
            'encrassement': [
                'Nettoyage chimique',
                'Filtres améliorés',
                'Rinçage périodique',
                'Contrôle pH auto',
                'Nettoyage par soufflage',
                'Analyse des dépôts'
            ],
            'vibration': [
                'Amortisseurs',
                'Contrôle serrage',
                'Analyse fréquentielle',
                'Équilibrage',
                'Modification des supports',
                'Renforts structurels'
            ],
            'fuite': [
                'Remplacement joints',
                'Test pression',
                'Contrôle serrage',
                'Analyse vibratoire',
                'Modification des raccords',
                'Inspection thermographique'
            ]
        }
        
        # Définir les seuils de criticité
        if criticality <= 12:
            action_count = 1  # Criticité négligeable
            maintenance_type = "Maintenance corrective"
        elif criticality <= 16:
            action_count = 2  # Criticité moyenne
            maintenance_type = "Maintenance préventive systématique"
        elif criticality <= 20:
            action_count = 2  # Criticité élevée
            maintenance_type = "Maintenance préventive conditionnelle"
        else:
            action_count = 3  # Criticité interdite
            maintenance_type = "Remise en cause complète"
        
        # Trouver la liste d'actions correspondante
        actions_list = []
        for cause_key, actions in base_actions.items():
            if cause_key in cause.lower():
                actions_list = actions
                break
        
        # Si aucune liste n'est trouvée, utiliser une liste générique
        if not actions_list:
            actions_list = [
                'Inspection régulière',
                'Maintenance préventive',
                'Analyse des causes',
                'Formation personnel',
                'Amélioration procédures',
                'Remplacement périodique'
            ]
        
        # S'assurer qu'on ne demande pas plus d'actions qu'il n'y en a de disponibles
        action_count = min(action_count, len(actions_list))
        
        # Sélectionner un sous-ensemble d'actions de manière aléatoire
        selected_actions = random.sample(actions_list, action_count)
        
        # Ajouter le type de maintenance au début
        selected_actions.insert(0, maintenance_type)
        
        # Joindre les actions avec des symboles '+'
        return " + ".join(selected_actions)
    
    def save_to_file(self, output_path=None):
        """
        Sauvegarde l'AMDEC générée dans un fichier Excel formaté
        
        Args:
            output_path (str, optional): Chemin de sortie pour le fichier Excel.
                Si non fourni, un chemin par défaut sera utilisé.
        
        Returns:
            str: Chemin du fichier sauvegardé
        """
        if self.amdec_df is None:
            raise ValueError("Aucune AMDEC à sauvegarder. Appelez d'abord la méthode generate().")
        
        if output_path is None:
            # Créer un répertoire models s'il n'existe pas
            models_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'models')
            os.makedirs(models_dir, exist_ok=True)
            
            # Créer un nom de fichier
            output_path = os.path.join(models_dir, "amdec_generated.xlsx")
        
        # Sauvegarder d'abord avec pandas
        self.amdec_df.to_excel(output_path, index=False)
        
        # Vérifier si le fichier a été correctement sauvegardé
        try:
            # Ensuite, formater avec openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Définir les styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Bordures
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Appliquer les styles à l'en-tête
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Ajuster la largeur des colonnes
            column_widths = {
                'A': 20,  # Composant
                'B': 20,  # Sous-composant
                'C': 20,  # Fonction
                'D': 25,  # Mode de Défaillance
                'E': 20,  # Cause
                'F': 20,  # Effet
                'G': 5,   # F
                'H': 5,   # G
                'I': 5,   # D
                'J': 5,   # C
                'K': 40   # Actions Correctives
            }
            
            for col, width in column_widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width
            
            # Appliquer un formatage conditionnel pour la criticité
            for row in range(2, len(self.amdec_df) + 2):
                criticality_cell_ref = f'J{row}'
                if criticality_cell_ref in ws:
                    criticality_cell = ws[criticality_cell_ref]
                    criticality_value = criticality_cell.value
                    
                    if criticality_value is not None:
                        # Convertir en entier si nécessaire
                        if isinstance(criticality_value, str):
                            try:
                                criticality_value = int(criticality_value)
                            except ValueError:
                                criticality_value = 0
                        
                        # Appliquer une couleur en fonction de la criticité
                        if criticality_value <= 12:
                            criticality_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Vert
                        elif criticality_value <= 16:
                            criticality_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Jaune
                        elif criticality_value <= 20:
                            criticality_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
                        else:
                            criticality_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
                
                # Ajouter des bordures et de l'alignement à toutes les cellules
                for col in range(1, 12):  # A à K
                    cell_ref = ws.cell(row=row, column=col).coordinate
                    if cell_ref in ws:
                        cell = ws[cell_ref]
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            # Sauvegarder le fichier formaté
            wb.save(output_path)
        except Exception as e:
            print(f"Attention: Erreur lors du formatage du fichier Excel: {str(e)}")
            print("Le fichier AMDEC a été sauvegardé sans formatage.")
        
        self.output_path = output_path
        return output_path
    
    def get_component_criticality(self, component, subcomponent):
        """
        Récupère la criticité pour un composant et sous-composant spécifiques
        
        Args:
            component (str): Nom du composant
            subcomponent (str): Nom du sous-composant
            
        Returns:
            int: Criticité ou 0 si non trouvée
        """
        if self.amdec_df is None:
            raise ValueError("Aucune AMDEC générée. Appelez d'abord la méthode generate().")
        
        # Normaliser les noms pour la recherche
        component = component.lower()
        subcomponent = subcomponent.lower()
        
        # Convertir les colonnes en minuscules pour la recherche
        try:
            component_col = self.amdec_df['Composant'].str.lower()
            subcomponent_col = self.amdec_df['Sous-composant'].str.lower()
            
            # Filtre pour trouver le composant et sous-composant
            filtered_df = self.amdec_df[
                (component_col == component) & 
                (subcomponent_col == subcomponent)
            ]
            
            if filtered_df.empty:
                return 0
            
            # Retourner la criticité maximale
            return filtered_df['C'].max()
        except Exception:
            # En cas d'erreur, retourner 0
            return 0